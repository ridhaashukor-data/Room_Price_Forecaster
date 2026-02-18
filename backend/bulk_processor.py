"""
Bulk Occupancy Forecasting - Excel Input/Output Processor

Handles:
1. Excel template generation
2. Parsing uploaded Excel files
3. Bulk forecasting for multiple dates
4. Generating Excel output with occupancy grid
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Color
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, Rule
import os

# Import core forecasting functions
from forecaster import (
    load_completion_ratios,
    forecast_occupancy,
    parse_date,
    CONFIG
)

# ============================================================================
# CONFIGURATION
# ============================================================================

BULK_CONFIG = {
    'max_forecast_days': 30,  # Only forecast up to n=30
    'template_filename': 'occupancy_template.xlsx',
    'output_filename_prefix': 'forecast_output'
}

# Conditional formatting colors (Excel hex colors)
COLORS = {
    'white_0':    'FFFFFFFF',    # 0% - white
    'green_50':   'FF92D050',    # 50% - green
    'yellow_75':  'FFFFFF00',    # 75% - yellow
    'red_99':     'FFFF0000',    # 99% - red
    'purple_100': 'FF9966FF',    # 100+ - purple
}

# ============================================================================
# EXCEL TEMPLATE GENERATOR
# ============================================================================

def generate_template(output_dir='./'):
    """
    Generate Excel template for bulk occupancy forecasting.
    
    Template structure (matches output format):
    - Upload Date
    - Grid: Date/Month | Jan | Jan_Forecast | Feb | Feb_Forecast | ...
    
    Returns: Path to generated template file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Occupancy Template"
    
    # ========================================================================
    # SECTION 1: UPLOAD DATE
    # ========================================================================
    
    # Header
    ws['A1'] = 'OCCUPANCY FORECASTING - INPUT TEMPLATE'
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = 'Upload Date (DD/MM/YY):'
    ws['B3'] = datetime.now().strftime('%d/%m/%y')
    ws['B3'].font = Font(italic=True)
    
    # ========================================================================
    # SECTION 2: OCCUPANCY GRID (Current + Forecast columns)
    # ========================================================================
    
    ws['A5'] = 'OCCUPANCY DATA (%)'
    ws['A5'].font = Font(bold=True)
    
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    # Border styles for grouping
    thin = Side(style='thin', color='FF000000')
    left_border            = Border(left=thin)
    right_border           = Border(right=thin)
    left_top_border        = Border(left=thin, top=thin)
    top_right_border       = Border(top=thin, right=thin)
    bottom_right_border    = Border(right=thin, bottom=thin)
    left_bottom_border     = Border(left=thin, bottom=thin)
    
    # Days in each month for 2026
    days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 2026 is not leap year
    
    # Header row (row 7): Date/Month | Jan | Jan_Forecast | Feb | Feb_Forecast | ...
    ws['A7'] = 'Date/Month'
    ws['A7'].font = Font(bold=True)
    
    col_idx = 2
    for month in months:
        # Current occupancy column
        ws.cell(row=7, column=col_idx, value=month)
        ws.cell(row=7, column=col_idx).font = Font(bold=True)
        
        # Forecast column (leave empty for template)
        ws.cell(row=7, column=col_idx+1, value=f'{month}_Forecast')
        ws.cell(row=7, column=col_idx+1).font = Font(bold=True, color='808080')
        
        # Top edge of grouped month-pair box
        ws.cell(row=7, column=col_idx).border = left_top_border
        ws.cell(row=7, column=col_idx+1).border = top_right_border
        
        col_idx += 2
    
    # Data rows: 1-31 for each date
    for date_num in range(1, 32):
        row_num = 7 + date_num
        ws.cell(row=row_num, column=1, value=date_num)
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        
        col_idx = 2
        for month_idx, month in enumerate(months):
            # Current occupancy (user fills)
            ws.cell(row=row_num, column=col_idx, value=0)
            
            # Forecast column (empty for template)
            ws.cell(row=row_num, column=col_idx+1, value=None)
            
            col_idx += 2
    
    # ========================================================================
    # CONDITIONAL FORMATTING (applied before borders)
    # ========================================================================
    # <50: no formatting (blocker rule with no style)
    ws.conditional_formatting.add(
        'B8:Y38',
        Rule(type='expression', formula=['B8<50'], stopIfTrue=True)
    )

    # >=100: purple
    ws.conditional_formatting.add('B8:Y38',
        CellIsRule(operator='greaterThanOrEqual', formula=['100'],
                   fill=PatternFill(start_color=COLORS['purple_100'], end_color=COLORS['purple_100'], fill_type='solid'),
                   stopIfTrue=True))
    # 3-color gradient: 50(green) -> 75(yellow) -> 99(red)
    ws.conditional_formatting.add('B8:Y38',
        ColorScaleRule(start_type='num', start_value=50,  start_color=Color(rgb=COLORS['green_50']),
                       mid_type='num',   mid_value=75,   mid_color=Color(rgb=COLORS['yellow_75']),
                       end_type='num',   end_value=99,   end_color=Color(rgb=COLORS['red_99'])))

    # ========================================================================
    # BORDERS (applied after conditional formatting)
    # Left border on current-occ col + right border on forecast col = grouped box
    # Bottom border at last valid day of each month
    # ========================================================================
    for date_num in range(1, 32):
        row_num = 7 + date_num
        col_idx = 2
        for month_idx in range(len(months)):
            cell_current  = ws.cell(row=row_num, column=col_idx)
            cell_forecast = ws.cell(row=row_num, column=col_idx + 1)
            last_day = date_num == days_in_month[month_idx]

            cell_current.border  = left_bottom_border  if last_day else left_border
            cell_forecast.border = bottom_right_border if last_day else right_border

            col_idx += 2
    
    # ========================================================================
    # INSTRUCTIONS
    # ========================================================================
    
    ws['A40'] = 'INSTRUCTIONS:'
    ws['A40'].font = Font(bold=True)
    ws['A41'] = '1. Update Upload Date to current date (DD/MM/YY format)'
    ws['A42'] = '2. Fill ONLY the month columns (not forecast columns)'
    ws['A43'] = '3. Enter occupancy 0-100% for each date'
    ws['A44'] = '4. Bottom borders mark last valid day of each month - do not enter data below these'
    ws['A45'] = '5. Upload to get forecast - forecast columns will be filled automatically'
    ws['A46'] = '6. Colors: <50 no color | 50 green, 75 yellow, 99 red | 100+ purple'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    for col in range(2, 26):  # Now have 24 columns (12 months x 2)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 10
    
    # Save template
    template_path = os.path.join(output_dir, BULK_CONFIG['template_filename'])
    wb.save(template_path)
    
    print(f"✅ Template generated: {template_path}")
    return template_path

# ============================================================================
# EXCEL PARSER
# ============================================================================

def parse_uploaded_excel(filepath):
    """
    Parse uploaded Excel file and extract inputs.
    
    Returns: dict with:
        - upload_date: datetime object
        - occupancy_df: DataFrame with columns [stay_date, current_occupancy]
    """
    print(f"📂 Parsing uploaded file: {filepath}")
    
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Extract upload date (B3)
    upload_date_str = ws['B3'].value
    if isinstance(upload_date_str, datetime):
        upload_date = upload_date_str
    else:
        upload_date = datetime.strptime(upload_date_str, '%d/%m/%y')
    
    # Extract current occupancy grid (rows 8-38, columns with current data only)
    # New structure: Col 2=Jan, 3=Jan_Forecast, 4=Feb, 5=Feb_Forecast, etc.
    # We read columns 2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24 (current only)
    occupancy_data = []
    
    current_year = upload_date.year
    month_nums = list(range(1, 13))  # 1-12
    
    for date_num in range(1, 32):
        row_idx = 7 + date_num
        
        for idx, month_num in enumerate(month_nums):
            # Read from current columns only (2, 4, 6, ..., 24)
            col_idx = 2 + (idx * 2)  # Even columns: 2, 4, 6, 8, etc.
            occupancy_value = ws.cell(row=row_idx, column=col_idx).value
            
            if occupancy_value is None:
                occupancy_value = 0
            else:
                occupancy_value = float(occupancy_value)
            
            # Create date object
            try:
                stay_date = datetime(current_year, month_num, date_num)
                occupancy_data.append({
                    'stay_date': stay_date,
                    'current_occupancy': occupancy_value
                })
            except ValueError:
                # Invalid date (e.g., Feb 31) - skip
                continue
    
    occupancy_df = pd.DataFrame(occupancy_data)
    
    print(f"✅ Parsed successfully:")
    print(f"   Upload date: {upload_date.strftime('%d/%m/%Y')}")
    print(f"   Occupancy records: {len(occupancy_df)}")
    
    return {
        'upload_date': upload_date,
        'occupancy_df': occupancy_df
    }

# ============================================================================
# BULK FORECASTING ENGINE
# ============================================================================

def bulk_forecast(parsed_inputs, completion_ratios_df=None):
    """
    Run forecasting for all dates in the occupancy DataFrame.
    
    Only forecasts for dates where:
    - stay_date >= upload_date (not in the past)
    - days_out <= 30 (within forecast window)
    - current_occupancy > 0 (has data)
    
    Returns: DataFrame with forecast results (occupancy only, no pricing)
    """
    print("\n🔄 Running bulk forecast...")
    
    # Load completion ratios if not provided
    if completion_ratios_df is None:
        completion_ratios_df = load_completion_ratios()
    
    upload_date = parsed_inputs['upload_date']
    occupancy_df = parsed_inputs['occupancy_df']
    
    # Use default values for required parameters
    default_total_rooms = 100
    
    results = []
    
    for idx, row in occupancy_df.iterrows():
        stay_date = row['stay_date']
        current_occ = row['current_occupancy']
        
        # Calculate days out
        days_out = (stay_date - upload_date).days
        
        # Skip past dates
        if days_out < 0:
            continue
        
        # Skip dates beyond forecast window
        if days_out > BULK_CONFIG['max_forecast_days']:
            continue
        
        # Skip if zero occupancy (no data to forecast)
        if current_occ == 0:
            continue
        
        # Build input for this specific date
        stay_date_str = stay_date.strftime('%d%m%y')
        upload_date_str = upload_date.strftime('%d%m%y')
        
        inputs = {
            'stay_date': stay_date_str,
            'today_date': upload_date_str,
            'current_occupancy': current_occ,
            'total_rooms_available': default_total_rooms,
            'event_level': 'none'  # Fixed for bulk mode
        }
        
        try:
            # Run forecast only (no pricing)
            forecast_results = forecast_occupancy(inputs, completion_ratios_df)
            
            # Combine results (occupancy data only)
            result = {
                'stay_date': stay_date,
                'current_occupancy': current_occ,
                **forecast_results
            }
            
            results.append(result)
            
        except Exception as e:
            print(f"⚠️  Error processing {stay_date.strftime('%d/%m/%Y')}: {e}")
            continue
    
    results_df = pd.DataFrame(results)
    
    print(f"✅ Bulk forecast complete: {len(results_df)} dates processed")
    
    return results_df

# ============================================================================
# OUTPUT GENERATOR - SINGLE SHEET EXCEL
# ============================================================================

def generate_output_excel(parsed_inputs, forecast_df, output_dir='./'):
    """
    Generate output Excel with occupancy grid:
    
    Shows current and forecast occupancy in calendar format
    
    Returns: Path to generated file
    """
    print("\n📊 Generating output Excel...")
    
    upload_date = parsed_inputs['upload_date']
    occupancy_df = parsed_inputs['occupancy_df']
    
    # Create workbook and match template layout exactly
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Occupancy Forecast"

    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    # Border styles
    thin = Side(style='thin', color='FF000000')
    left_border = Border(left=thin)
    right_border = Border(right=thin)
    left_top_border = Border(left=thin, top=thin)
    top_right_border = Border(top=thin, right=thin)
    left_bottom_border = Border(left=thin, bottom=thin)
    bottom_right_border = Border(right=thin, bottom=thin)

    # Header block (same as template)
    ws1['A1'] = 'OCCUPANCY FORECASTING - RESULTS'
    ws1['A1'].font = Font(bold=True, size=14)

    ws1['A3'] = 'Upload Date (DD/MM/YY):'
    ws1['B3'] = upload_date.strftime('%d/%m/%y')
    ws1['B3'].font = Font(italic=True)

    ws1['A5'] = 'OCCUPANCY DATA (%)'
    ws1['A5'].font = Font(bold=True)

    ws1['A7'] = 'Date/Month'
    ws1['A7'].font = Font(bold=True)

    # Fast lookups for current and forecast values
    current_lookup = {}
    for _, row in occupancy_df.iterrows():
        current_lookup[(row['stay_date'].month, row['stay_date'].day)] = row['current_occupancy']

    forecast_lookup = {}
    for _, row in forecast_df.iterrows():
        forecast_lookup[(row['stay_date'].month, row['stay_date'].day)] = row['forecast_occupancy_pct']

    # Month headers (row 7)
    col_idx = 2
    for month in months:
        ws1.cell(row=7, column=col_idx, value=month)
        ws1.cell(row=7, column=col_idx).font = Font(bold=True)

        ws1.cell(row=7, column=col_idx + 1, value=f'{month}_Forecast')
        ws1.cell(row=7, column=col_idx + 1).font = Font(bold=True, color='808080')

        # Top outside border for grouped month pair
        ws1.cell(row=7, column=col_idx).border = left_top_border
        ws1.cell(row=7, column=col_idx + 1).border = top_right_border

        col_idx += 2

    # Data grid (rows 8-38)
    for date_num in range(1, 32):
        row_num = 7 + date_num
        ws1.cell(row=row_num, column=1, value=date_num)
        ws1.cell(row=row_num, column=1).font = Font(bold=True)

        col_idx = 2
        for month_idx in range(12):
            month_num = month_idx + 1
            key = (month_num, date_num)

            try:
                datetime(upload_date.year, month_num, date_num)
            except ValueError:
                ws1.cell(row=row_num, column=col_idx, value=None)
                ws1.cell(row=row_num, column=col_idx + 1, value=None)
                col_idx += 2
                continue

            current_occ = current_lookup.get(key, 0)
            forecast_occ = forecast_lookup.get(key, None)

            ws1.cell(row=row_num, column=col_idx, value=current_occ)
            forecast_cell = ws1.cell(row=row_num, column=col_idx + 1)
            if forecast_occ is None:
                forecast_cell.value = None
            else:
                forecast_cell.value = round(float(forecast_occ), 1)
                forecast_cell.number_format = '0.0'

            col_idx += 2

    # ========================================================================
    # CONDITIONAL FORMATTING (same range and rules as template)
    # ========================================================================
    ws1.conditional_formatting.add(
        'B8:Y38',
        Rule(type='expression', formula=['B8<50'], stopIfTrue=True)
    )

    ws1.conditional_formatting.add('B8:Y38',
        CellIsRule(operator='greaterThanOrEqual', formula=['100'],
                   fill=PatternFill(start_color=COLORS['purple_100'], end_color=COLORS['purple_100'], fill_type='solid'),
                   stopIfTrue=True))

    ws1.conditional_formatting.add('B8:Y38',
        ColorScaleRule(start_type='num', start_value=50, start_color=Color(rgb=COLORS['green_50']),
                       mid_type='num', mid_value=75, mid_color=Color(rgb=COLORS['yellow_75']),
                       end_type='num', end_value=99, end_color=Color(rgb=COLORS['red_99'])))

    # ========================================================================
    # BORDERS: Outside grouping per month pair, plus valid-day bottom borders
    # ========================================================================
    for date_num in range(1, 32):
        row_num = 7 + date_num
        col_idx = 2
        for month_idx in range(12):
            cell_current = ws1.cell(row=row_num, column=col_idx)
            cell_forecast = ws1.cell(row=row_num, column=col_idx + 1)
            last_day = date_num == days_in_month[month_idx]

            cell_current.border = left_bottom_border if last_day else left_border
            cell_forecast.border = bottom_right_border if last_day else right_border

            col_idx += 2

    # Instructions block (same as template)
    ws1['A40'] = 'INSTRUCTIONS:'
    ws1['A40'].font = Font(bold=True)
    ws1['A41'] = '1. Update Upload Date to current date (DD/MM/YY format)'
    ws1['A42'] = '2. Fill ONLY the month columns (not forecast columns)'
    ws1['A43'] = '3. Enter occupancy 0-100% for each date'
    ws1['A44'] = '4. Bottom borders mark last valid day of each month - do not enter data below these'
    ws1['A45'] = '5. Upload to get forecast - forecast columns will be filled automatically'
    ws1['A46'] = '6. Colors: <50 no color | 50 green, 75 yellow, 99 red | 100+ purple'

    # Adjust column widths (same as template)
    ws1.column_dimensions['A'].width = 15
    for col in range(2, 26):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = 10
    
    # ========================================================================
    # SAVE FILE
    # ========================================================================
    
    output_filename = f"{BULK_CONFIG['output_filename_prefix']}_{upload_date.strftime('%Y%m%d')}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    wb.save(output_path)
    
    print(f"✅ Output generated: {output_path}")
    print(f"   Occupancy data: 31 rows")
    
    return output_path

# ============================================================================
# MAIN BULK PROCESSING FUNCTION
# ============================================================================

def process_bulk_forecast(input_excel_path, output_dir='./', completion_ratios_df=None):
    """
    Complete bulk forecasting workflow:
    1. Parse input Excel
    2. Run forecasts
    3. Generate output Excel
    
    Returns: Path to output file
    """
    print("="*70)
    print("🏨 BULK OCCUPANCY FORECASTING")
    print("="*70)
    
    # Parse input
    parsed_inputs = parse_uploaded_excel(input_excel_path)
    
    # Run forecasts
    forecast_df = bulk_forecast(parsed_inputs, completion_ratios_df=completion_ratios_df)
    
    # Generate output
    output_path = generate_output_excel(parsed_inputs, forecast_df, output_dir)
    
    print("\n" + "="*70)
    print("✅ BULK PROCESSING COMPLETE")
    print("="*70)
    
    return output_path

# ============================================================================
# EXAMPLE USAGE
# ============================================================================

if __name__ == "__main__":
    # Generate template
    print("Generating Excel template...")
    template_path = generate_template()
    
    print("\n" + "="*70)
    print("📝 Next steps:")
    print("1. Open the generated template")
    print("2. Fill in your occupancy data")
    print("3. Run process_bulk_forecast() with your filled template")
    print("="*70)
